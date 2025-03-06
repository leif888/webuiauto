import os

from openpyxl import load_workbook

from util_tools.logs_util.recordlog import logs


class ExcelDataReader:
    """
    读取Excel文件数据
    """

    def __init__(self, file_path):
        self.file_path = file_path
        abspath = os.path.abspath(self.file_path)

        if not os.path.isfile(abspath):
            raise FileNotFoundError(f'文件路径不存在：{abspath}')

        # read_only：控制工作簿是否以只读模式打开，如果为True，工作簿以只读模式打开，不允许对工作簿进行修改
        self.workbook = load_workbook(self.file_path, read_only=False)

    def read_entire_row(self, sheet_name='Sheet1', row_index=1):
        """
        获取Excel文件一整行的数据
        :param sheet_name: sheet页名称
        :param row_index: 要返回哪一行的数据，索引从1开始
        :return:
        """
        try:
            sheet = self.workbook[sheet_name]
            rows_data = [cell.value for cell in sheet[row_index]]
            return rows_data
        except Exception as e:
            logs.error(f'读取Excel一整列数据异常，原因为：{e}')
        finally:
            self.close()

    def read_multiple_rows(self, sheet_name='Sheet1'):
        """
        获取Excel多行的数据，每一行数据组成一个列表
        :param sheet_name: sheet页名称
        :return: 返回list，如：[['admin123','123456'],[],[],...]
        """
        try:
            all_rows_data = []
            sheet = self.workbook[sheet_name]
            max_col = sheet.max_column  # 获取到Excel的最大列数
            # min_row：从哪一行开始读取；max_row：获取最大行数；min_col：从哪一列开始读取；max_col：获取最大列数
            for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=max_col):
                row_data = [cell.value for cell in row]
                all_rows_data.append(row_data)
            return all_rows_data
        except Exception as e:
            logs.error(f'读取所有行数据异常，原因为：{e}')
        finally:
            self.close()

    def read_entire_column(self, sheet_name='Sheet1', column_index=1):
        """
        获取Excel文件一整列的数据
        :param sheet_name: sheet页名称
        :param column_index: 要返回哪一列的数据，索引从1开始
        :return:
        """
        columns_data = []
        try:
            sheet = self.workbook[sheet_name]
            for row in sheet.iter_rows(min_row=1, min_col=column_index, max_col=column_index):
                columns_data.append(row[0].value)
            return columns_data
        except Exception as e:
            logs.error(f'读取整列数据异常，原因为：{e}')
        finally:
            self.close()

    def read_cell_value(self, row_index, column_index, sheet_name='Sheet1'):
        """
        获取单元格的数据
        :param sheet_name: sheet页名称
        :param row_index: 行索引，从1开始
        :param column_index: 列索引，从1开始
        :return:
        """
        try:
            sheet = self.workbook[sheet_name]
            return sheet.cell(row=row_index, column=column_index).value
        except Exception as e:
            logs.error(f'读取Excel单元格数据异常，原因为：{e}')
        finally:
            self.close()

    def close(self):
        self.workbook.close()
